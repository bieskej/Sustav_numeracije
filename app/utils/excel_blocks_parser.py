from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class AssignedBlock:
    operator_name: str
    msisdn_od: int
    msisdn_do: int
    tip: str
    ndc: int | None = None
    blok: int | None = None
    nsn_len: int | None = None
    additional_info: str | None = None


def _norm_header(v: object) -> str:
    return str(v or "").strip().casefold()


def parse_assigned_blocks_xlsx(path: Path) -> list[AssignedBlock]:
    """
    Deterministic parser for your regulator XLSX structure (BIH).

    Expected columns (case-insensitive, partial match):
    - NDC: 'nacionalni odredišni kod', 'ndc'
    - blok: 'blok brojeva', 'blok'
    - N(S)N length: 'n(s)n', 'dužina broja', 'duzina broja', 'min', 'max'
    - operator: 'telekom operator', 'operator'
    - additional: 'dodatne informacije'

    Range mapping rule (provided by you):
    - N(S)N length is typically 8 digits
    - full N(S)N is constructed as: NDC + blok + padding digits to reach nsn_len
      Example: 30 + 3020 + XX => 30302000..30302099

    We import only assigned geographic blocks: operator != 'Nije dodijeljen'.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    h = [_norm_header(x) for x in header]

    def find_col(*needles: str) -> int | None:
        for i, hv in enumerate(h):
            for n in needles:
                if n in hv:
                    return i + 1
        return None

    col_ndc = find_col("nacionalni odredi", "ndc")
    col_blok = find_col("blok brojeva", "blok")
    col_operator = find_col("telekom operator", "operator")
    col_add = find_col("dodatne informacije")
    col_min = find_col("min")
    col_max = find_col("max")
    col_nsn = find_col("n(s)n", "dužina broja", "duzina broja")

    if not (col_ndc and col_blok and col_operator and (col_nsn or (col_min and col_max))):
        raise ValueError(
            f"Ne mogu pronaći potrebne kolone u XLSX headeru: {header}"
        )

    blocks: list[AssignedBlock] = []
    for r in range(2, ws.max_row + 1):
        operator = " ".join(str(ws.cell(row=r, column=col_operator).value or "").split())
        if not operator or operator.casefold() == "nije dodijeljen":
            continue

        ndc_raw = ws.cell(row=r, column=col_ndc).value
        blok_raw = ws.cell(row=r, column=col_blok).value
        if ndc_raw is None or blok_raw is None:
            continue

        try:
            ndc = int(str(ndc_raw).strip())
            blok = int(str(blok_raw).strip())
        except ValueError:
            continue

        # N(S)N length: prefer explicit N(S)N column, otherwise use min/max (must match)
        nsn_len: int | None = None
        if col_nsn:
            v = ws.cell(row=r, column=col_nsn).value
            if v is not None and str(v).strip() != "":
                try:
                    nsn_len = int(str(v).strip())
                except ValueError:
                    nsn_len = None
        if nsn_len is None and col_min and col_max:
            vmin = ws.cell(row=r, column=col_min).value
            vmax = ws.cell(row=r, column=col_max).value
            try:
                vmin_i = int(str(vmin).strip())
                vmax_i = int(str(vmax).strip())
            except Exception:
                continue
            if vmin_i != vmax_i:
                continue
            nsn_len = vmin_i

        if nsn_len is None or nsn_len <= 0:
            continue

        prefix = f"{ndc}{blok}"
        pad = nsn_len - len(prefix)
        if pad < 0:
            continue

        od = int(prefix + ("0" * pad))
        do = int(prefix + ("9" * pad))

        additional_info = None
        if col_add:
            additional_info = " ".join(str(ws.cell(row=r, column=col_add).value or "").split()) or None

        blocks.append(
            AssignedBlock(
                operator_name=operator,
                msisdn_od=od,
                msisdn_do=do,
                tip="geografski",
                ndc=ndc,
                blok=blok,
                nsn_len=nsn_len,
                additional_info=additional_info,
            )
        )

    return blocks

